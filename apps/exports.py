"""
Excel va PDF eksport funksiyalari.
Bu faylni apps/exports.py sifatida saqlang.
"""
import io
from datetime import date

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import Order, Product, Cliente, Payment

# ════════════════════════════════════════════════════════════
# UMUMIY YORDAMCHI FUNKSIYALAR
# ════════════════════════════════════════════════════════════

BRAND_COLOR = "534AB7"  # Excel uchun hex (# siz)
BRAND_COLOR_HEX = "#534AB7"  # PDF (reportlab) uchun

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color=BRAND_COLOR, end_color=BRAND_COLOR)
TITLE_FONT = Font(name='Calibri', bold=True, size=16, color=BRAND_COLOR)
CELL_FONT = Font(name='Calibri', size=10.5)
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)


def _style_header_row(ws, row_num, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _style_data_row(ws, row_num, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center')


def _auto_width(ws, ncols, min_width=12, max_width=42):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 3, max_width))


def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pdf_response(buf, filename):
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _pdf_table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(BRAND_COLOR_HEX)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F7')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ])


def _pdf_header(story, title, subtitle=None):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleBrand', parent=styles['Title'],
        textColor=colors.HexColor(BRAND_COLOR_HEX), fontSize=18, spaceAfter=4,
    )
    story.append(Paragraph(title, title_style))
    if subtitle:
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], textColor=colors.grey, fontSize=9)
        story.append(Paragraph(subtitle, sub_style))
    story.append(Spacer(1, 12))


def _apply_order_filters(request):
    """order_list bilan bir xil filtr mantig'ini eksport uchun qayta ishlatish."""
    q = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    orders = Order.objects.select_related('cliente', 'agent').all()
    if q:
        orders = orders.filter(
            Q(number__icontains=q) |
            Q(cliente__first_name__icontains=q) |
            Q(cliente__firma_name__icontains=q) |
            Q(agent__first_name__icontains=q)
        )
    if status_filter:
        orders = orders.filter(status=status_filter)
    if date_from:
        orders = orders.filter(date_created__date__gte=date_from)
    if date_to:
        orders = orders.filter(date_created__date__lte=date_to)
    return orders.order_by('-date_created')


def _apply_product_filters(request):
    q = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    products = Product.objects.filter(is_active=True)
    if q:
        products = products.filter(Q(name__icontains=q) | Q(sku__icontains=q))
    if status_filter == 'low':
        products = products.filter(stock__gt=0, stock__lte=10)
    elif status_filter == 'out':
        products = products.filter(stock=0)
    return products


def _apply_cliente_filters(request):
    q = request.GET.get('q', '')
    clientes = Cliente.objects.filter(is_active=True).select_related('agent')
    if q:
        clientes = clientes.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(firma_name__icontains=q) |
            Q(alternative_name__icontains=q) |
            Q(phone__icontains=q)
        )
    return clientes


def _apply_payment_filters(request):
    """payment_list bilan bir xil filtr mantig'ini eksport uchun qayta ishlatish."""
    method_filter = request.GET.get('method', '')
    confirmed_filter = request.GET.get('confirmed', '')
    q = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    payments = Payment.objects.select_related('order__cliente', 'order__agent').all()

    if method_filter:
        payments = payments.filter(method=method_filter)
    if confirmed_filter == '1':
        payments = payments.filter(confirmed=True)
    elif confirmed_filter == '0':
        payments = payments.filter(confirmed=False)
    if q:
        payments = payments.filter(
            Q(order__number__icontains=q) |
            Q(order__cliente__first_name__icontains=q) |
            Q(order__cliente__last_name__icontains=q) |
            Q(order__cliente__firma_name__icontains=q) |
            Q(order__cliente__alternative_name__icontains=q)
        )
    if date_from:
        payments = payments.filter(date_created__date__gte=date_from)
    if date_to:
        payments = payments.filter(date_created__date__lte=date_to)

    return payments.order_by('-date_created')


# ════════════════════════════════════════════════════════════
# NAKLADNOYLAR — EXCEL
# ════════════════════════════════════════════════════════════

@login_required
def order_export_excel(request):
    orders = _apply_order_filters(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nakladnoylar"

    headers = ['Raqam', 'Sana', 'Mijoz', 'Firma', 'Agent', 'To\'lov turi', 'Jami summa', 'Holat']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    status_labels = dict(Order.Status.choices)
    payment_labels = dict(Order.PaymentType.choices)

    row_num = 2
    for o in orders:
        ws.append([
            o.number,
            o.date_created.strftime('%d.%m.%Y %H:%M'),
            f"{o.cliente.first_name} {o.cliente.last_name}",
            o.cliente.firma_name or '—',
            f"{o.agent.first_name} {o.agent.last_name}",
            payment_labels.get(o.payment_type, o.payment_type),
            o.total_sum,
            status_labels.get(o.status, o.status),
        ])
        _style_data_row(ws, row_num, len(headers))
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        row_num += 1

    # Jami summa qatori
    ws.append(['', '', '', '', '', 'JAMI:', f'=SUM(G2:G{row_num - 1})', ''])
    total_row = row_num
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0'

    _auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'

    filename = f"nakladnoylar_{date.today().strftime('%Y%m%d')}.xlsx"
    return _excel_response(wb, filename)


# ════════════════════════════════════════════════════════════
# NAKLADNOYLAR — PDF
# ════════════════════════════════════════════════════════════

@login_required
def order_export_pdf(request):
    orders = _apply_order_filters(request)
    status_labels = dict(Order.Status.choices)
    payment_labels = dict(Order.PaymentType.choices)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=12 * mm, rightMargin=12 * mm,
    )
    story = []
    _pdf_header(story, "Nakladnoylar ro'yxati",
                f"Yaratilgan sana: {date.today().strftime('%d.%m.%Y')} | Jami: {orders.count()} ta")

    data = [['Raqam', 'Sana', 'Mijoz', 'Agent', "To'lov turi", 'Summa', 'Holat']]
    total = 0
    for o in orders:
        data.append([
            o.number,
            o.date_created.strftime('%d.%m.%Y'),
            f"{o.cliente.first_name} {o.cliente.last_name}"[:28],
            f"{o.agent.first_name} {o.agent.last_name}"[:22],
            payment_labels.get(o.payment_type, o.payment_type),
            f"{o.total_sum:,.0f}",
            status_labels.get(o.status, o.status),
        ])
        total += o.total_sum

    data.append(['', '', '', '', 'JAMI:', f"{total:,.0f}", ''])

    table = Table(data, repeatRows=1, colWidths=[60, 55, 130, 110, 75, 75, 70])
    style = _pdf_table_style()
    style.add('FONTNAME', (4, -1), (5, -1), 'Helvetica-Bold')
    table.setStyle(style)
    story.append(table)

    doc.build(story)
    buf.seek(0)
    filename = f"nakladnoylar_{date.today().strftime('%Y%m%d')}.pdf"
    return _pdf_response(buf, filename)


# ════════════════════════════════════════════════════════════
# TOVARLAR — EXCEL
# ════════════════════════════════════════════════════════════

@login_required
def product_export_excel(request):
    products = _apply_product_filters(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tovarlar"

    headers = ['Nomi', 'SKU', 'Narx (so\'m)', 'Skladda', 'Ogohlantirish chegarasi', 'Holat']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    status_map = {'ok': 'Yetarli', 'low': 'Kam qolgan', 'out': 'Tugagan'}

    row_num = 2
    for p in products:
        ws.append([
            p.name, p.sku, p.price, p.stock, p.low_stock_threshold,
            status_map.get(p.stock_status, p.stock_status),
        ])
        _style_data_row(ws, row_num, len(headers))
        ws.cell(row=row_num, column=3).number_format = '#,##0'
        # Kam/tugagan holatlarni rangli belgilash
        if p.stock_status == 'out':
            ws.cell(row=row_num, column=6).font = Font(color='A32D2D', bold=True)
        elif p.stock_status == 'low':
            ws.cell(row=row_num, column=6).font = Font(color='854F0B', bold=True)
        row_num += 1

    _auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'

    filename = f"tovarlar_{date.today().strftime('%Y%m%d')}.xlsx"
    return _excel_response(wb, filename)


# ════════════════════════════════════════════════════════════
# TOVARLAR — PDF
# ════════════════════════════════════════════════════════════

@login_required
def product_export_pdf(request):
    products = _apply_product_filters(request)
    status_map = {'ok': 'Yetarli', 'low': 'Kam qolgan', 'out': 'Tugagan'}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=15 * mm, rightMargin=15 * mm,
    )
    story = []
    _pdf_header(story, "Tovarlar ro'yxati (Sklad)",
                f"Yaratilgan sana: {date.today().strftime('%d.%m.%Y')} | Jami: {products.count()} ta")

    data = [['Nomi', 'SKU', 'Narx', 'Miqdor', 'Holat']]
    for p in products:
        data.append([
            p.name[:32], p.sku, f"{p.price:,.0f}", str(p.stock),
            status_map.get(p.stock_status, p.stock_status),
        ])

    table = Table(data, repeatRows=1, colWidths=[160, 80, 75, 60, 90])
    table.setStyle(_pdf_table_style())
    story.append(table)

    doc.build(story)
    buf.seek(0)
    filename = f"tovarlar_{date.today().strftime('%Y%m%d')}.pdf"
    return _pdf_response(buf, filename)


# ════════════════════════════════════════════════════════════
# MIJOZLAR — EXCEL
# ════════════════════════════════════════════════════════════

@login_required
def cliente_export_excel(request):
    clientes = _apply_cliente_filters(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"

    headers = ['Ism', 'Familiya', 'Firma', 'Alternativ nom', 'Telefon', 'Manzil', 'Agent', 'Holat']
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    row_num = 2
    for c in clientes:
        ws.append([
            c.first_name, c.last_name, c.firma_name or '—', c.alternative_name or '—',
                                       c.phone or '—', c.address or '—', str(c.agent) if c.agent else '—',
            'Faol' if c.is_active else 'Nofaol',
        ])
        _style_data_row(ws, row_num, len(headers))
        row_num += 1

    _auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'

    filename = f"mijozlar_{date.today().strftime('%Y%m%d')}.xlsx"
    return _excel_response(wb, filename)


# ════════════════════════════════════════════════════════════
# MIJOZLAR — PDF
# ════════════════════════════════════════════════════════════

@login_required
def cliente_export_pdf(request):
    clientes = _apply_cliente_filters(request)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=12 * mm, rightMargin=12 * mm,
    )
    story = []
    _pdf_header(story, "Mijozlar ro'yxati",
                f"Yaratilgan sana: {date.today().strftime('%d.%m.%Y')} | Jami: {clientes.count()} ta")

    data = [['Ism Familiya', 'Firma', 'Alternativ nom', 'Telefon', 'Agent', 'Holat']]
    for c in clientes:
        data.append([
            f"{c.first_name} {c.last_name}"[:28],
            (c.firma_name or '—')[:24],
            (c.alternative_name or '—')[:24],
            c.phone or '—',
            str(c.agent) if c.agent else '—',
            'Faol' if c.is_active else 'Nofaol',
        ])

    table = Table(data, repeatRows=1, colWidths=[140, 110, 110, 85, 95, 65])
    table.setStyle(_pdf_table_style())
    story.append(table)

    doc.build(story)
    buf.seek(0)
    filename = f"mijozlar_{date.today().strftime('%Y%m%d')}.pdf"
    return _pdf_response(buf, filename)


# ════════════════════════════════════════════════════════════
# TO'LOVLAR — EXCEL
# ════════════════════════════════════════════════════════════

@login_required
def payment_export_excel(request):
    payments = _apply_payment_filters(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "To'lovlar"

    headers = [
        'Sana', 'Nakladnoy', 'Mijoz', 'Firma', 'Alternativ nom',
        'Agent', 'Summa', 'Turi', 'Holat',
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    row_num = 2
    for p in payments:
        ws.append([
            p.date_created.strftime('%d.%m.%Y %H:%M'),
            p.order.number,
            f"{p.order.cliente.first_name} {p.order.cliente.last_name}",
            p.order.cliente.firma_name or '—',
            p.order.cliente.alternative_name or '—',
            f"{p.order.agent.first_name} {p.order.agent.last_name}",
            p.amount,
            'Naqd' if p.method == 'cash' else 'Bank',
            'Tasdiqlangan' if p.confirmed else 'Kutilmoqda',
        ])
        _style_data_row(ws, row_num, len(headers))
        ws.cell(row=row_num, column=7).number_format = '#,##0'
        if not p.confirmed:
            ws.cell(row=row_num, column=9).font = Font(color='854F0B', bold=True)
        row_num += 1

    # Jami summa qatori
    ws.append(['', '', '', '', '', 'JAMI:', f'=SUM(G2:G{row_num - 1})', '', ''])
    total_row = row_num
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0'

    _auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'

    filename = f"tolovlar_{date.today().strftime('%Y%m%d')}.xlsx"
    return _excel_response(wb, filename)


# ════════════════════════════════════════════════════════════
# TO'LOVLAR — PDF
# ════════════════════════════════════════════════════════════

@login_required
def payment_export_pdf(request):
    payments = _apply_payment_filters(request)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=15 * mm, bottomMargin=15 * mm, leftMargin=12 * mm, rightMargin=12 * mm,
    )
    story = []
    total_amount = payments.aggregate(s=Sum('amount'))['s'] or 0
    _pdf_header(
        story, "To'lovlar ro'yxati",
        f"Yaratilgan sana: {date.today().strftime('%d.%m.%Y')} | Jami: {payments.count()} ta | Jami summa: {total_amount:,.0f} so'm"
    )

    data = [['Sana', 'Nakladnoy', 'Mijoz', 'Alternativ nom', 'Agent', 'Summa', 'Turi', 'Holat']]
    for p in payments:
        data.append([
            p.date_created.strftime('%d.%m.%Y'),
            p.order.number,
            f"{p.order.cliente.first_name} {p.order.cliente.last_name}"[:24],
            (p.order.cliente.alternative_name or '—')[:20],
            f"{p.order.agent.first_name} {p.order.agent.last_name}"[:20],
            f"{p.amount:,.0f}",
            'Naqd' if p.method == 'cash' else 'Bank',
            'Tasdiqlangan' if p.confirmed else 'Kutilmoqda',
        ])

    data.append(['', '', '', '', 'JAMI:', f"{total_amount:,.0f}", '', ''])

    table = Table(data, repeatRows=1, colWidths=[60, 60, 110, 95, 95, 75, 55, 80])
    style = _pdf_table_style()
    style.add('FONTNAME', (4, -1), (5, -1), 'Helvetica-Bold')
    table.setStyle(style)
    story.append(table)

    doc.build(story)
    buf.seek(0)
    filename = f"tolovlar_{date.today().strftime('%Y%m%d')}.pdf"
    return _pdf_response(buf, filename)
